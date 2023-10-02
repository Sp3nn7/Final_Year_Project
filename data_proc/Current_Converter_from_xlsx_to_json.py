import json
import openpyxl

def converter(file_name):
    wb = openpyxl.load_workbook(file_name)
    all_data = []
    states = []
    itemName = []
    counter = 0
    regions = []
    categories = []
    for state in wb:
        print(state)
        current_state_data = []
        states.append(str(state).split('"')[1])
        regions.append([state.cell(row=1, column=i).value for i in range(2, state.max_column+1)])
        for k in range(2, 7940):
            if counter == 0:
                itemName.append(state.cell(row=k, column=1).value)
                categories.append(state.cell(row=k, column=2).value)
            dummy = [state.cell(row=k, column=i).value for i in range(3, state.max_column+1)]
            data = []
            for x in dummy:
                if x != None:
                    data.append(x)
                else:
                    data.append(x)
            current_state_data.append(data)
        counter += 1
        all_data.append(current_state_data)

    dictionary = {
        "data":
            all_data,
        "datamap": {
            "state": states,
            "itemName": itemName,
            "category": categories,
            "region": regions
        }
    }

    json_object = json.dumps(dictionary, indent=4)
    with open("../src/data/current_data.json", "w") as outfile:
        outfile.write(json_object)


converter('FoldedRegions_new.xlsx')
