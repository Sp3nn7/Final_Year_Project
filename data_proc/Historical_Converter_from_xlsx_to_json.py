import json
import openpyxl


def converter(file_name):
    wb = openpyxl.load_workbook(file_name)
    all_data = []
    cities = []
    categories = []
    dates = []
    inflation = []
    counter = 0
    unit_size = []
    for city in wb:
        current_city_data = []
        cities.append(str(city).split('"')[1])
        for k in range(3, city.max_row+1):
            if counter == 0:
                categories.append(city.cell(row=k, column=1).value)
                dates = [city.cell(row=2, column=i).value for i in range(3, 113)]
                inflation = [city.cell(row=1, column=i).value for i in range(3, 113)]
                unit_size = [city.cell(row=i, column=2).value for i in range(3, 113)]
            dummy = [city.cell(row=k, column=i).value for i in range(3, 113)]
            data = []
            for x in dummy:
                if x is not None:
                    data.append(int(x) / 100)
                else:
                    data.append(x)
            current_city_data.append(data)
        counter += 1
        all_data.append(current_city_data)
    for index_date in range(len(dates)):
        if dates[index_date] is not None:
            dates[index_date] = dates[index_date].strftime('%Y-%m-%d')

    dictionary = {
        "data":
            all_data,
        "datamap": {
            "city": cities,
            "category": categories,
            "unit": unit_size,
            "date": dates,
            "inflation": inflation
        }
    }

    json_object = json.dumps(dictionary, indent=4)
    with open("historical_data.json", "w") as outfile:
        outfile.write(json_object)


converter('Collated.xlsx')
